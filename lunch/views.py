# from datetime import date
import calendar
import json
import io
import os
from datetime import date, time, timedelta
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from weasyprint import HTML
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.conf import settings
from openpyxl.cell.cell import MergedCell

from .models import Order


def get_allowed_dates(start: date, count: int) -> set[date]:
    """
    start日から「日曜を除く」count 日間を集めてセットで返す。
    """
    allowed = set()
    d = start
    # count 分だけ日曜以外をcollect
    while len(allowed) < count:
        if d.weekday() != 6:  # 6=日曜
            allowed.add(d)
        d += timedelta(days=1)
    return allowed

@login_required
def monthly_calendar(request, year=None, month=None):
    """月間カレンダー表示ビュー"""
    today = date.today()
    year  = year or today.year
    month = month or today.month

    # ── 追加: 許可日を計算 ──
    allowed_dates = get_allowed_dates(today, 6)
    # テンプレートでも today と allowed_dates を参照できるように渡す

    cal = calendar.Calendar(firstweekday=0)
    month_days = cal.monthdatescalendar(year, month)

    # ユーザーの今月の注文日をセット化
    orders = Order.objects.filter(
        user=request.user,
        order_date__year=year,
        order_date__month=month,
        canceled=False
    ).values_list('order_date', flat=True)
    orders_set = set(orders)

    # 日付ごとに「注文済みか」を付与
    calendar_data = []
    for week in month_days:
        week_data = []
        for day in week:
            week_data.append({
                'day': day,
                'is_current_month': day.month == month,
                'ordered': day in orders_set,
                'allowed': day in allowed_dates,
            })
        calendar_data.append(week_data)

    return render(request, 'lunch/calendar.html', {
        'calendar_data': calendar_data,
        'year': year,
        'month': month,
        'today': today,
        'allowed_dates': allowed_dates,
    })
def fax_order_pdf(request):
    # 今日は何件注文があるか集計
    today = date.today()
    qs = Order.objects.filter(order_date=today, canceled=False)
    # ライスの大中小それぞれ注文数を仮に集計する例
    counts = {
        'large':  qs.filter(rice_size='大').count(),
        'medium': qs.filter(rice_size='中').count(),
        'small':  qs.filter(rice_size='小').count(),
    }

    # HTML をレンダリングして PDF 化
    html_string = render_to_string('lunch/fax_template.html', {
        'today': today,
        'counts': counts,
    })
    html = HTML(string=html_string)
    pdf = html.write_pdf()

    # レスポンスで返却
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = today.strftime('lunch_order_%Y%m%d.pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def today_order(request):
    """当日の注文を行う／キャンセルする画面"""
    user = request.user
    today = date.today()

    # まず、当日の注文レコードをキャンセルフラグに関係なく取得
    order = Order.objects.filter(user=user, order_date=today).first()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'order':
            if not order:
                # レコード自体がなければ新規作成
                Order.objects.create(user=user, order_date=today, vendor='veg17', rice_size='中')
            else:
                # 既存レコードがあるならキャンセルフラグをオフに
                order.canceled = False
                order.canceled_at = None
                order.save()
        elif action == 'cancel' and order:
            # キャンセルするときはフラグをオンにして日時を埋める
            order.canceled = True
            order.canceled_at = timezone.now()
            order.save()
        return redirect('today_order')

    # ★キャンセル済みの場合はテンプレート上では「注文なし扱い」にする
    order_for_template = order if order and not order.canceled else None
    return render(request, 'lunch/today_order.html', {
        'order': order_for_template,
        'today': today,
    })
@login_required
@require_POST
def toggle_order(request):
    """
    POST JSON { "date": "YYYY-MM-DD" }
    → その日の注文レコードを必ず取得 or 作成し、canceled フラグをトグル
    """
    data = json.loads(request.body)
    try:
        day = date.fromisoformat(data['date'])
    except Exception:
        return JsonResponse({'error': 'invalid date'}, status=400)

    # 当日9:00以降は締め切り
    now = timezone.localtime()
    # 当日から日曜除く6日以外は変更不可
    allowed = get_allowed_dates(date.today(), 6)
    if day not in allowed:
        return JsonResponse({'error': '変更可能期間外です'}, status=403)

    if now.time() >= time(8, 10) and day == date.today():
        return JsonResponse({'error': '受付は午前9時までです'}, status=403)
    # get_or_create ならレコードがなければ作ってくれる
    order, created = Order.objects.get_or_create(
        user=request.user,
        order_date=day,
        defaults={'vendor': 'veg17', 'rice_size': '中'}
    )

    # 事務がステータスを発注済にした後は変更不可
    if hasattr(order, 'status') and order.status != 'pending':
        return JsonResponse({'error': '既に発注済のため変更できません'}, status=403)
    # トグル処理
    if order.canceled:
        order.canceled = False
        order.canceled_at = None
        status = 'ordered'
    else:
        order.canceled = True
        order.canceled_at = timezone.now()
        status = 'canceled'

    order.save()
    return JsonResponse({'status': status, 'date': data['date']})


@staff_member_required
def download_monthly_report(request, year=None, month=None):
    """
    管理者(staff)専用ビュー。
    GET パラメータ ?year=YYYY&month=MM があればそれを使い、なければ本日を基準に出力。
    レスポンスとして Excel ファイルを返却する。
    """
    # 1) リクエストから年月を取得（なければ本日）
    today = date.today()
    y = int(request.GET.get('year', year or today.year))
    m = int(request.GET.get('month', month or today.month))

    # 2) LunchConfig を取得 or デフォルト作成
    cfg, created = LunchConfig.objects.get_or_create(
        defaults={'price': 430, 'subsidy': 230, 'monthly_limit': 3780}
    )
    if not created and cfg.price != 430:
        # 既存レコードの単価が古いままなら上書き
        cfg.price = 430
        cfg.save()

    # 3) ユーザー一覧と日数を取得
    User = get_user_model()
    users = User.objects.all().order_by('username')
    days_in_month = calendar.monthrange(y, m)[1]

    # 4) Excel ワークブック／シートを組み立て
    wb = Workbook()
    ws = wb.active
    ws.title = f"{y}年{m}月ランチ注文"

    # 4-1) ヘッダー行
    header = ['コード','氏名'] \
           + [f"{d}日" for d in range(1, days_in_month+1)] \
           + ['注文数','合計金額','補助額','上限','会社負担','超過','実費']
    ws.append(header)
    # ヘッダー書式
    for col in range(1, len(header)+1):
        cell = ws.cell(row=1, column=col)
        cell.font      = Font(bold=True)
        cell.fill      = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(horizontal='center')

    # 4-2) 曜日行
    weekday_map = ['月','火','水','木','金','土','日']
    weekday_row = ['', '']
    for d in range(1, days_in_month+1):
        wd = date(y, m, d).weekday()  # 0=月 … 6=日
        weekday_row.append(weekday_map[wd])
    weekday_row += [''] * 7
    ws.append(weekday_row)
    for col in range(1, len(header)+1):
        cell = ws.cell(row=2, column=col)
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(italic=True)

    # 4-3) 週末列リストを作成（セル番号）
    weekend_cols = []
    for d in range(1, days_in_month+1):
        wd = date(y, m, d).weekday()  # 5=土, 6=日
        if wd in (5, 6):
            weekend_cols.append(2 + d)  # 「コード」「氏名」を飛ばして 3 列目から日付

    # 4-4) 各ユーザー行を挿入
    for row_idx, user in enumerate(users, start=3):
        code = user.id
        name = user.get_full_name() or user.username

        # 日別フラグ (1 or 0)
        flags = [
            1 if Order.objects.filter(
                    user=user,
                    order_date=date(y, m, d),
                    canceled=False
               ).exists() else 0
            for d in range(1, days_in_month+1)
        ]

        # 集計列を計算
        total_qty     = sum(flags)
        total_price   = total_qty * cfg.price
        total_subsidy = total_qty * cfg.subsidy
        limit         = cfg.monthly_limit
        company_pay   = min(total_subsidy, limit)
        over          = max(0, total_subsidy - limit)
        user_pay      = total_price - company_pay

        row = [code, name] + flags + [
            total_qty,
            total_price,
            total_subsidy,
            limit,
            company_pay,
            over,
            user_pay,
        ]
        ws.append(row)

        # 4-4-1) 週末セルを灰色に
        for col in weekend_cols:
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = PatternFill("solid", fgColor="EEEEEE")

        # 4-4-2) 通貨列にカンマ区切り書式を設定
        base_col = 2 + days_in_month
        for offset in range(1, 7):  # 「合計金額」から「実費」までの 6 列
            cell = ws.cell(row=row_idx, column=base_col + offset)
            cell.number_format = '"¥"#,##0'

    # 4-5) 日別合計行を追加
    total_row = ['', '合計']
    # 日別合計を計算
    daily_totals = []
    for d in range(1, days_in_month+1):
        cnt = Order.objects.filter(
            order_date=date(y, m, d),
            canceled=False
        ).count()
        daily_totals.append(cnt)
    total_qty     = sum(daily_totals)
    total_price   = total_qty * cfg.price
    total_subsidy = total_qty * cfg.subsidy
    limit         = cfg.monthly_limit
    company_pay   = min(total_subsidy, limit)
    over          = max(0, total_subsidy - limit)
    user_pay      = total_price - company_pay

    total_row += daily_totals + [
        total_qty,
        total_price,
        total_subsidy,
        limit,
        company_pay,
        over,
        user_pay,
    ]
    ws.append(total_row)

    # 合計行の週末セルを灰色に
    total_row_idx = ws.max_row
    for col in weekend_cols:
        cell = ws.cell(row=total_row_idx, column=col)
        cell.fill = PatternFill("solid", fgColor="EEEEEE")

    # 合計行の書式（太字＆黄色背景）
    for col in range(1, len(header)+1):
        cell = ws.cell(row=total_row_idx, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFFF99")

    # 4-6) ベンダー集計行をシート下部に追加
    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1, value='≪ベンダー集計≫').font = Font(bold=True)
    for i, (code, name) in enumerate(Order.VENDORS, start=start_row+1):
        qs = Order.objects.filter(
            order_date__year=y,
            order_date__month=m,
            vendor=code,
            canceled=False
        )
        cnt = qs.count()
        amt = sum(o.price for o in qs)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=days_in_month+3, value=cnt)
        ws.cell(row=i, column=days_in_month+4, value=amt)

    # 5) Workbook をバイト列に書き込み、HttpResponse で返す
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'lunch_report_{y}{m:02}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # ブラウザ側でダウンロードさせるヘッダー
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return 

@login_required
def fax_order_excel(request):
    today = date.today()
    qs = Order.objects.filter(order_date=today, canceled=False)
    counts = {
        '大': qs.filter(rice_size='大').count(),
        '中': qs.filter(rice_size='中').count(),
        '小': qs.filter(rice_size='小').count(),
    }

    template_path = os.path.join(
        settings.BASE_DIR, 'lunch', 'templates', 'lunch', 'fax_template.xlsx'
    )
    wb = load_workbook(template_path)
    ws = wb.active

    # 書き込みたいセルと値のマッピング
    writes = {
        # 年月日を個別セルに
        'B11': today.year,      # 年
        'D11': today.month,     # 月
        'F11': today.day,       # 日
        'H11': counts['大'],                  # ライス大
        'J11': counts['中'],                  # ライス中
        'L11': counts['小'],                  # ライス小
    }

    for coord, val in writes.items():
        cell = ws[coord]  # 例: ws['E6']
        # MergedCell ならスキップ（またはマージ範囲の左上セルを優先）
        if isinstance(cell, MergedCell):
            # もしマージ範囲の左上セルを書き込みたい場合は、
            # 下記のように範囲を探して min_row/min_col で再設定してください。
            for m in ws.merged_cells.ranges:
                if coord in m:
                    tl = (m.min_row, m.min_col)
                    ws.cell(row=tl[0], column=tl[1]).value = val
                    break
        else:
            cell.value = val

    # 出力
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = today.strftime('lunch_order_%Y%m%d.xlsx')
    r = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    r['Content-Disposition'] = f'attachment; filename="{filename}"'
    return r
# def fax_order_excel(request):
    
#     """当日の注文を紙イメージに近いフォーマットの Excel でダウンロード"""
#     today = date.today()
#     qs = Order.objects.filter(order_date=today, canceled=False)
#     counts = {
#         '大': qs.filter(rice_size='大').count(),
#         '中': qs.filter(rice_size='中').count(),
#         '小': qs.filter(rice_size='小').count(),
#     }

#     # Workbook／Sheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = '注文書'

#     # スタイル定義
#     thin = Side(border_style="thin", color="000000")
#     border = Border(top=thin, bottom=thin, left=thin, right=thin)
#     bold = Font(bold=True)
#     center = Alignment(horizontal='center', vertical='center')

#     # 1) ヘッダー部
#     ws.merge_cells('A1:D1')
#     h1 = ws['A1']
#     h1.value = 'ベジタブルディッシュ１７　御中　FAX:0191‐43‐2121'
#     h1.font = Font(size=12)
#     h1.alignment = center

#     ws.merge_cells('A2:D2')
#     h2 = ws['A2']
#     h2.value = 'お弁当注文書'
#     h2.font = Font(bold=True, size=14)
#     h2.alignment = center

#     # 2) 注文集計テーブル
#     # 見出し行
#     ws.append(['注文日', 'ライス大', 'ライス中', 'ライス小'])
#     for col in range(1,5):
#         c = ws.cell(row=3, column=col)
#         c.font = bold
#         c.alignment = center
#         c.border = border

#     # データ行
#     ws.append([
#         f"{today.year}年{today.month}月{today.day}日",
#         counts['大'],
#         counts['中'],
#         counts['小'],
#     ])
#     for col in range(1,5):
#         c = ws.cell(row=4, column=col)
#         c.alignment = center
#         c.border = border

#     # 3) 確認印欄
#     ws.merge_cells('A6:B6')
#     ws['A6'].value = '確認印'
#     ws['A6'].font = bold
#     ws['A6'].alignment = center
#     ws['A6'].border = border
#     # 印欄の空白セルにも枠を
#     c = ws['C6']
#     c.border = border
#     c = ws['D6']
#     c.border = border

#     # 4) フッター部（右下に会社情報）
#     ws.merge_cells('C8:D8')
#     f1 = ws['C8']
#     f1.value = 'NSエンジニアリング株式会社'
#     f1.font = Font(size=10)
#     f1.alignment = Alignment(horizontal='right')
#     ws.merge_cells('C9:D9')
#     f2 = ws['C9']
#     f2.value = 'FAX:0191‐48‐4491'
#     f2.font = Font(size=10)
#     f2.alignment = Alignment(horizontal='right')
#     ws.merge_cells('C10:D10')
#     f3 = ws['C10']
#     f3.value = '※捺印にて返送お願いいたします'
#     f3.font = Font(size=10)
#     f3.alignment = Alignment(horizontal='right')

#     # 列幅の調整（お好みで微調整してください）
#     ws.column_dimensions['A'].width = 20
#     ws.column_dimensions['B'].width = 12
#     ws.column_dimensions['C'].width = 12
#     ws.column_dimensions['D'].width = 12

#     # 出力
#     output = io.BytesIO()
#     wb.save(output)
#     output.seek(0)

#     filename = today.strftime('lunch_order_%Y%m%d.xlsx')
#     response = HttpResponse(
#         output.read(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#     return response
