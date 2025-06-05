import calendar
from datetime import date, datetime
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from lunch.models import Order, LunchConfig
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers

class Command(BaseCommand):
    help = "月末ランチ注文レポートを Excel で出力します"

    def add_arguments(self, parser):
        parser.add_argument('--year',  type=int, default=date.today().year)
        parser.add_argument('--month', type=int, default=date.today().month)

    def handle(self, *args, **options):
        # ── LunchConfig の取得 or 作成 ──
        cfg, created = LunchConfig.objects.get_or_create(
            defaults={'price': 430, 'subsidy': 200, 'monthly_limit': 3780}
        )
        if created:
            self.stdout.write(self.style.WARNING(
                'LunchConfig レコードが存在しなかったため、デフォルト値で自動作成しました'
            ))

        year  = options['year']
        month = options['month']
        User  = get_user_model()
        days  = calendar.monthrange(year, month)[1]
        users = User.objects.all().order_by('username')

        # ワークブック＆シート
        wb  = Workbook()
        ws  = wb.active
        ws.title = f"{year}年{month}月ランチ注文"

        # ヘッダー行（コード, 氏名, 1日～N日, 集計列）
        header = ['コード','氏名'] \
               + [f"{d}日" for d in range(1, days+1)] \
               + ['注文数計','合計金額','補助額','上限','会社負担','超過','実費']
        ws.append(header)

        # ── 曜日行を追加 ──
        # headerの日付部分に対応する曜日（日本語：月〜日）を生成
        weekday_map = ['月','火','水','木','金','土','日']
        weekday_row = ['', '']  # 「コード」「氏名」列は空白
        for d in range(1, days+1):
            wd = date(year, month, d).weekday()  # 0=月 … 6=日
            weekday_row.append(weekday_map[wd])
        # 集計列の曜日は空白にしておく
        weekday_row += [''] * 7
        ws.append(weekday_row)

        # 曜日行の書式（中央寄せ＆イタリックなど）
        for col in range(1, len(header)+1):
            cell = ws.cell(row=2, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(italic=True)

        # ヘッダー書式
        for col in range(1, len(header)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
            cell.alignment = Alignment(horizontal='center')
        # ── 週末列を灰色にするための列インデックスリスト作成 ──
        weekend_cols = []
        for d in range(1, days+1):
            wd = datetime(year, month, d).weekday()  # 5=土,6=日
            if wd in (5, 6):
                # 「コード」「氏名」を飛ばして、日付列は3列目から始まる
                weekend_cols.append(2 + d)

        # 各ユーザー行
        for row_idx, user in enumerate(users, start=2):
            code = user.id
            name = user.get_full_name() or user.username

            # 日別フラグ
            flags = [
                1 if Order.objects.filter(
                        user=user,
                        order_date=date(year, month, d),
                        canceled=False
                   ).exists() else 0
                for d in range(1, days+1)
            ]

            # 集計
            total_qty     = sum(flags)
            total_price   = total_qty * cfg.price
            total_subsidy = total_qty * cfg.subsidy
            limit         = cfg.monthly_limit
            company_pay   = min(total_subsidy, limit)
            over          = max(0, total_subsidy - limit)
            user_pay      = total_price - company_pay

            row = [code, name] + flags + [
                total_qty,
                total_price,
                total_subsidy,
                limit,
                company_pay,
                over,
                user_pay,
            ]
            ws.append(row)

            # 通貨列にカンマ区切り書式を設定
            # 「注文数」が列 index = 2 + days
            base_col = 2 + days
            for offset, fmt_col in enumerate(
                ['合計金額','補助額','上限','会社負担','超過','実費'], start=1
            ):
                cell = ws.cell(
                    row=row_idx,
                    column=base_col + offset
                )
                # ¥付きカンマ区切り
                cell.number_format = '"¥"#,##0'

            # ── データ行の週末セルを灰色に ──
            for col in weekend_cols:
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = PatternFill("solid", fgColor="EEEEEE")

        # 日別合計行
        total_row = ['', '合計']
        # 日別合計
        for d in range(1, days+1):
            cnt = Order.objects.filter(
                order_date=date(year, month, d),
                canceled=False
            ).count()
            total_row.append(cnt)
        # 合計列集計
        total_qty     = sum(total_row[2:2+days])
        total_price   = total_qty * cfg.price
        total_subsidy = total_qty * cfg.subsidy
        limit         = cfg.monthly_limit
        company_pay   = min(total_subsidy, limit)
        over          = max(0, total_subsidy - limit)
        user_pay      = total_price - company_pay
        total_row += [
            total_qty,
            total_price,
            total_subsidy,
            limit,
            company_pay,
            over,
            user_pay,
        ]
        ws.append(total_row)

        last = ws.max_row
        for offset in range(1, 7):  # 集計列は 7 列分
            cell = ws.cell(
                row=last,
                column=base_col + offset
            )
            cell.number_format = '"¥"#,##0'

        # ── 合計行の週末セルも灰色に ──
        total_row_idx = ws.max_row
        for col in weekend_cols:
            cell = ws.cell(row=total_row_idx, column=col)
            cell.fill = PatternFill("solid", fgColor="EEEEEE")

        # 合計行書式
        last_row = ws.max_row
        for col in range(1, len(header)+1):
            cell = ws.cell(row=last_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFFF99")

        # 下部にベンダー集計
        start = ws.max_row + 2
        ws.cell(row=start, column=1, value='≪ベンダー集計≫').font = Font(bold=True)
        for i, (code, name) in enumerate(Order.VENDORS, start= start+1):
            qs = Order.objects.filter(
                order_date__year=year,
                order_date__month=month,
                vendor=code,
                canceled=False
            )
            cnt = qs.count()
            amt = sum(o.price for o in qs)
            ws.cell(row=i, column=2, value=name)
            ws.cell(row=i, column=days+3, value=cnt)
            ws.cell(row=i, column=days+4, value=amt)

        # 保存
        filename = f'lunch_report_{year}{month:02}.xlsx'
        wb.save(filename)
        self.stdout.write(self.style.SUCCESS(f'レポートを {filename} に出力しました'))
