import calendar
from datetime import date
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from lunch.models import Order, LunchConfig
import openpyxl
from openpyxl import Workbook

class Command(BaseCommand):
    help = "月末ランチ注文レポートを Excel で出力します"

    def add_arguments(self, parser):
        parser.add_argument('--year',  type=int, default=date.today().year)
        parser.add_argument('--month', type=int, default=date.today().month)

    def handle(self, *args, **options):
        cfg, created = LunchConfig.objects.get_or_create(
            defaults={'price': 430, 'subsidy': 200, 'monthly_limit': 3780}
        )
        if created:
            self.stdout.write(self.style.WARNING(
                'LunchConfig レコードが存在しなかったため、デフォルト値で自動作成しました'
            ))
        
        year  = options['year']
        month = options['month']
        User  = get_user_model()
        days  = calendar.monthrange(year, month)[1]
        users = User.objects.all().order_by('username')

        # Excel ブック／シート準備
        wb  = Workbook()
        ws1 = wb.active
        if ws1 is None:
            ws1 = wb.create_sheet()
        ws1.title = '日別注文'

        # ヘッダー行: コード, 氏名, 1日〜N日, 注文数, 合計金額, 補助額, 上限, 会社負担, 超過, 実費
        header = ['コード', '氏名'] \
               + [f"{d}日" for d in range(1, days+1)] \
               + ['注文数','合計金額','補助額','上限','会社負担','超過','実費']
        ws1.append(header)
        # 各ユーザーごとに1行ずつ
        for user in users:
            # 社員コードは user.id または別途 field があればそちらを
            code = user.id  
            name = user.get_full_name() or user.username

            # 日別フラグを取得
            flags = [
                1 if Order.objects.filter(
                        user=user,
                        order_date=date(year, month, d),
                        canceled=False
                   ).exists()
                  else 0
                for d in range(1, days+1)
            ]

            # 集計値を計算
            total_qty     = sum(flags)
            total_price   = total_qty * cfg.price
            total_subsidy = total_qty * cfg.subsidy
            limit         = cfg.monthly_limit
            company_pay   = min(total_subsidy, limit)
            over          = max(0, total_subsidy - limit)
            user_pay      = total_price - company_pay

            row = [code, name] + flags \
                + [ total_qty, total_price,
                    total_subsidy, limit,
                    company_pay, over, user_pay ]
            ws1.append(row)

            # ── ここから日別合計行を追加 ──
            total_row = ['', '合計']
            # 各日ごとの合計（0/1 のフラグをカウント）
            for d in range(1, days+1):
                cnt = Order.objects.filter(
                    order_date=date(year, month, d),
                    canceled=False
                ).count()
                total_row.append(cnt)
            # まとめ列の計算
            total_qty     = sum(total_row[2:2+days])
            total_price   = total_qty * cfg.price
            total_subsidy = total_qty * cfg.subsidy
            limit         = cfg.monthly_limit
            company_pay   = min(total_subsidy, limit)
            over          = max(0, total_subsidy - limit)
            user_pay      = total_price - company_pay
            total_row += [
                total_qty,
                total_price,
                total_subsidy,
                limit,
                company_pay,
                over,
                user_pay,
            ]
            ws1.append(total_row)
            # ───────────────────────────────────

        # （任意）下部にベンダー別集計行を追加
        start_row = ws1.max_row + 2
        ws1.cell(row=start_row, column=1, value='≪ベンダー集計≫')
        for i, (code,name) in enumerate(Order.VENDORS, start=1):
            qs_v = Order.objects.filter(
                order_date__year=year,
                order_date__month=month,
                vendor=code,
                canceled=False
            )
            cnt = qs_v.count()
            amt = sum(o.price for o in qs_v)
            ws1.append(['', name, *['']*(days-1), cnt, amt])

        # ファイル保存
        filename = f'lunch_report_{year}{month:02}.xlsx'
        wb.save(filename)
        self.stdout.write(self.style.SUCCESS(f'レポートを {filename} として出力しました'))

        # 1行目ヘッダー: 社員 + 日付
        # days = calendar.monthrange(year, month)[1]
        # headers = ['社員'] + [f"{d}日" for d in range(1, days+1)]
        # ws1.append(headers)



        # 各ユーザーの日別注文(1 or 0)を埋める
        # for user in users:
        #     row = [user.get_full_name()]
        #     for d in range(1, days+1):
        #         cnt = Order.objects.filter(
        #             user=user,
        #             order_date=date(year, month, d),
        #             canceled=False
        #         ).exists()
        #         row.append(1 if cnt else 0)
        #     ws1.append(row)

        # # 集計シート
        # ws2 = wb.create_sheet('集計')
        # ws2.append(['社員','注文数','合計金額','補助額','会社負担','超過','実費'])
        # for user in users:
        #     qs       = Order.objects.filter(
        #         user=user,
        #         order_date__year=year,
        #         order_date__month=month,
        #         canceled=False
        #     )
        #     total_qty    = qs.count()
        #     total_price  = sum(o.price   for o in qs)
        #     total_subsidy= sum(o.subsidy for o in qs)
        #     company_pay  = min(total_subsidy, cfg.monthly_limit)
        #     over         = max(0, total_subsidy - cfg.monthly_limit)
        #     user_pay     = total_price - company_pay
        #     ws2.append([
        #         user.get_full_name(),
        #         total_qty,
        #         total_price,
        #         total_subsidy,
        #         company_pay,
        #         over,
        #         user_pay,
        #     ])

        # # ベンダー集計シート(オプション)
        # ws3 = wb.create_sheet('ベンダー集計')
        # ws3.append(['ベンダー','注文数','合計金額'])
        # for code, name in Order.VENDORS:
        #     qs_v = Order.objects.filter(
        #         order_date__year
        #         =year,
        #         order_date__month=month,
        #         vendor=code,
        #         canceled=False
        #     )
        #     ws3.append([
        #         name,
        #         qs_v.count(),
        #         sum(o.price for o in qs_v),
        #     ])

        # # ファイル保存
        # filename = f'lunch_report_{year}{month:02}.xlsx'
        # wb.save(filename)
        # self.stdout.write(self.style.SUCCESS(f'レポートを {filename} に出力しました'))
